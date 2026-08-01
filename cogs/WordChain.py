import discord
from discord.ext import commands
import fcts.sqlcontrol as q
import fcts.etcfunctions as etc
import fcts.i18n_runtime as i18n
import fcts.leaderboard as l
import requests
import re
from fcts.koreanbreak import count_break_korean
import fcts.worddict as wd
import fcts.koreansearch as ks
import datetime
import random
import asyncio
from project_paths import DATA_DIR
from fcts.user_resolver import UserResolutionError, resolve_discord_user


def _read_workbook_rows(path):
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book.worksheets[0]
        return list(sheet.iter_rows(values_only=True))
    finally:
        book.close()


async def _prepare_wordchain_participant(ctx, reference=None):
    """Resolve a participant and create missing game dependencies safely."""
    user = await resolve_discord_user(ctx, reference)
    if getattr(user, "bot", False):
        raise UserResolutionError("Bots cannot participate in word chain games.")

    user_id = int(user.id)
    if not q.accountExistsById(user_id):
        q.newAccountById(user_id, user.name)
    if not q.storageExistsById(user_id):
        q.newStorageById(user_id)

    stats = l.wcReadById(user_id, "stats")
    if stats is None:
        l.wcUpdateRegistById(user_id)
        stats = l.wcReadById(user_id, "stats")

    return user, stats

player_badge = [
    "",
    "<:0_1:1294660411081228328>","<:0_2:1294660419625160734>",
    "<:0_3:1294660428097392712>","<:0_4:1294660438423900262>",
    "<:0_5:1294660447265357854>","<:0_6:1294660457646522410>",
    "<:0_7:1294660471282204733>","<:0_8:1294660480580845590>"
]

color = [0, 0xFFB3BA, 0xFFDFBA, 0xFFFFBA, 0xBAFFC9, 0xBAE1FF, 0xEECBFF, 0xFFD4E5, 0xA39193]

def scoreFont(value, digits, type=0):
    num=[
        [
            "<:255_0:1349993274160451636>",
            "<:255_1:1349993283409022986>",
            "<:255_2:1349993291210555462>",
            "<:255_3:1349993298801987626>",
            "<:255_4:1349993307362562088>",
            "<:255_5:1349993315675799592>",
            "<:255_6:1349993323959423007>",
            "<:255_7:1349993331882459188>",
            "<:255_8:1349993339964882944>",
            "<:255_9:1349993348009824279>",
            "<:255_10:1349993355932864512>",
            "<:255_11:1349993364459618306>"
        ],[
            "<:1_0:1349993396512624673>",
            "<:1_1:1349993403693404221>",
            "<:1_2:1349993412371415072>",
            "<:1_3:1349993419006808105>",
            "<:1_4:1349993426372005989>",
            "<:1_5:1349993434173407274>",
            "<:1_6:1349993441420902441>",
            "<:1_7:1349993449767567412>",
            "<:1_8:1349993456671658037>",
            "<:1_9:1349993464678584341>",
            "<:1_10:1349993471850577982>",
            "<:1_11:1349993478884429845>"
        ],[
            "<:2_0:1349993486094569572>",
            "<:2_1:1349993493547712553>",
            "<:2_2:1349993500715909161>",
            "<:2_3:1349993507842162728>",
            "<:2_4:1349993516310204416>",
            "<:2_5:1349993522333352006>",
            "<:2_6:1349993530742935562>",
            "<:2_7:1349993536753373317>",
            "<:2_8:1349993544214908948>",
            "<:2_9:1349993551290826812>",
            "<:2_10:1349993558572007464>",
            "<:2_11:1349993565735878666>"
        ],[
            "<:3_0:1349993572149100656>",
            "<:3_1:1349993580122472498>",
            "<:3_2:1349993587181486130>",
            "<:3_3:1349993595482144868>",
            "<:3_4:1349993607158956073>",
            "<:3_5:1349993614385741836>",
            "<:3_6:1349993622442868816>",
            "<:3_7:1349993629778837625>",
            "<:3_8:1349993637307482182>",
            "<:3_9:1349993647160164352>",
            "<:3_10:1349993655640916059>",
            "<:3_11:1349993663073222717>"
        ],[
            "<:4_0:1349993671252250624>",
            "<:4_1:1349993678759788616>",
            "<:4_2:1349993686703931432>",
            "<:4_3:1349993694845206528>",
            "<:4_4:1349993702533238816>",
            "<:4_5:1349993709919539202>",
            "<:4_6:1349993717427339326>",
            "<:4_7:1349993726390566952>",
            "<:4_8:1349993735106199672>",
            "<:4_9:1349993745495363649>",
            "<:4_10:1349993753062146100>",
            "<:4_11:1349993760942981233>"
        ],[
            "<:5_0:1349995193285218334>",
            "<:5_1:1349995201216905260>",
            "<:5_2:1349995210205302824>",
            "<:5_3:1349995217612312617>",
            "<:5_4:1349995225841401866>",
            "<:5_5:1349995233605193768>",
            "<:5_6:1349995241440018453>",
            "<:5_7:1349995249426235404>",
            "<:5_8:1349995257512857610>",
            "<:5_9:1349995265582436383>",
            "<:5_10:1349995273316864061>",
            "<:5_11:1349995282909102100>"
        ],[
            "<:6_0:1349995304107380778>",
            "<:6_1:1349995314957783083>",
            "<:6_2:1349995324139114496>",
            "<:6_3:1349995334842974329>",
            "<:6_4:1349995344942862368>",
            "<:6_5:1349995353620877384>",
            "<:6_6:1349995361472614441>",
            "<:6_7:1349995371706712105>",
            "<:6_8:1349995382741930035>",
            "<:6_9:1349995390736400395>",
            "<:6_10:1349995399481524254>",
            "<:6_11:1349995410189455412>"
        ],[
            "<:7_0:1349996843160961025>",
            "<:7_1:1349996852795277343>",
            "<:7_2:1349996861603319838>",
            "<:7_3:1349996873347239947>",
            "<:7_4:1349996881664802826>",
            "<:7_5:1349996890028249209>",
            "<:7_6:1349996899163443210>",
            "<:7_7:1349996907736596490>",
            "<:7_8:1349996915974209627>",
            "<:7_9:1349996924203175947>",
            "<:7_10:1349996934621958155>",
            "<:7_11:1349996943102840842>"
        ],[
            "<:8_0:1349996951592239124>",
            "<:8_1:1349996959313825844>",
            "<:8_2:1349996967270285323>",
            "<:8_3:1349996976602878014>",
            "<:8_4:1349996985528221696>",
            "<:8_5:1349996994424209481>",
            "<:8_6:1349997003379048488>",
            "<:8_7:1349997012287881247>",
            "<:8_8:1349997020546600990>",
            "<:8_9:1349997031686410303>",
            "<:8_10:1349997040662351953>",
            "<:8_11:1349997053882925107>"
        ]
    ]

    if digits < len(str(value)):
        digits = len(str(value))

    if value >= 0:
        temp = str(value)
        result = ""

        for i in range(digits - len(temp)):
            result += num[type][10]

        for char in temp:
            result += num[type][int(char)]

        return result
    
    else:
        temp = str(value)[1:]
        result = num[type][11]

        for i in range(digits - len(temp) - 1):
            result += num[type][10]

        for char in temp:
            result += num[type][int(char)]

        return result


def replace_sound_char(char):
    SOUND_LIST = {
        "라": "나",
        "락": "낙",
        "란": "난",
        "랄": "날",
        "람": "남",
        "랍": "납",
        "랑": "낭",
        "래": "내",
        "랭": "냉",
        "냑": "약",
        "략": "약",
        "냥": "양",
        "량": "양",
        "녀": "여",
        "려": "여",
        "녁": "역",
        "력": "역",
        "년": "연",
        "련": "연",
        "녈": "열",
        "렬": "열",
        "념": "염",
        "렴": "염",
        "렵": "엽",
        "녕": "영",
        "령": "영",
        "녜": "예",
        "례": "예",
        "로": "노",
        "록": "녹",
        "론": "논",
        "롱": "농",
        "뢰": "뇌",
        "뇨": "요",
        "료": "요",
        "룡": "용",
        "루": "누",
        "뉴": "유",
        "류": "유",
        "뉵": "육",
        "륙": "육",
        "륜": "윤",
        "률": "율",
        "륭": "융",
        "륵": "늑",
        "름": "늠",
        "릉": "능",
        "니": "이",
        "리": "이",
        "린": "인",
        "림": "임",
        "립": "입",
        "0": "영",
        "1": "일",
        "·": "점",
        "0": "영",
        "Ɩ": "일",
        "ς": "이",
        "Ɛ": "삼",
        "μ": "사",
        "ट": "오",
        "მ": "육",
        "٢": "칠",
        "8": "팔",
        "୧": "구",
        "✩": "별"
    }
    if char in SOUND_LIST:
        return SOUND_LIST[char]
    else:
        return None


def isOneKill(word):
    ONEKILL_WORD = [
        '겊', '귬', '깆', '껸', '꼇', '꼍', '꾜', '끠', '냔', '른', '늄', '랒', '읖',
        '릇', '쿄', '룅', '븀', '럴', '텝', '엌', '탉', '텝', '튬', '듐', '눞', '틤', '풂',
        '픔', '핕', '휵', '읗', '틋', '틂', '톹', '훽', '콫', '냘', '뇰', '뉼', '늉'
    ]
    if word in ONEKILL_WORD:
        return True
    else:
        return False


def searchKiller(start, length=0):
    HARD_WORD = [
        '겊', '귬', '깆', '껸', '꼇', '꼍', '꾜', '끠', '냔', '른', '녘', '늄', '랒', '읖',
        '릇', '쿄', '룅', '륨', '븀', '럴', '텝', '엌', '탉', '텝', '튬', '듐', '눞', '틤', '풂',
        '픔', '핕', '휵', '읗', '빱', '믄', '쁨', '궈', '뤄', '삸', '갊', '랏', '긔', '뮴', '틋',
        '틂', '톹', '훽', '콫', '냘', '뇰', '뉼', '늉', '덟', '돎', '듈', '랖', '랸', '럿', '렁',
        '렝', '롸', '룔', '륀', '릅', '릇', '릊', '릎', '먕', '믐', '밗', '볜', '븐', '븜', '앝',
        '엌', '왑', '웤', '읓', '읔', '읕', '읖', '읗', '잌', '쭘', '쭹', '웡', '찱', '캇', '쾃',
        '쿄', '탉', '텝', '곹', '궃', '궆', '궘', '긑', '깞', '꺠', '껱', '껼', '꽅', '꽌', '꾈',
        '꿑', '뀨', '낕', '넠', '녝', '녬', '놩', '뇸', '눤', '닼', '돍', '돜', '땽', '뚭', '뜹',
        '띱', '랓', '렃', '렄', '렆', '롕', '롶', '롹', '뢔', '뤌', '밲', '뤗', '릋', '릏', '먄',
        '멐', '볌', '봠', '븣', '붤', '븡', '풰', '빝', '뼌', '샄', '샆', '샡', '섳', '솣', '솦',
        '쇔', '숡', '솤', '싥', '싴', '썀', '쎂', '쎕', '얨', '얶', '옄', '왙', '욈', '웆', '웉',
        '윶', '읅', '읨', '쟤', '쟹', '젘', '짗', '캍', '컽', '쾜', '큭', '큿', '탘', '톔', '툿',
        '퓜', '훕', '픠', '윰', '쭝'
    ]
    result = []
    for char in HARD_WORD:
        result += wd.searchSpecial(start, char, length)

    return result

def detectZwong(index, player):
    max_i = len(player) - 1
    detect = None
    if index == max_i:
        detect = 0
    else:
        detect = index + 1

    if player[index]['id'] == 262899129276039169:
        print("Zwong Detected")
        return True
    else:
        print("사격중지 아군이다!")
        return False
    

#UI
def lifeUI(life, max):
    icon = [':black_heart:', ':heart:']
    return icon[1] * life + icon[0] * (max - life)


def scoreBoost(score, life):
    if life == 3:
        return score
    elif life == 2:
        return int(score * 1.16)
    elif life == 1:
        return int(score * 1.39)


def sampleText():
    sample = [
        "별을노래하는마음으로",
        "한송이의국화꽃을피우기위해",
        "가나다라마바사아자차카타파하",
        "내입술의말과",
        "희푸른모니터너머의빛을통해",
        "반짝이는재가당신의불꽃을따라",
        "생명가득한하늘을보여주자",
        "내희망의내용은질투뿐",
        "리갤말고리게로",
        "쌀독에서인심난다",
        "은비의비밀일기장",
        "영일이삼사오육칠팔구",
        "그저나답게빛나는거예요",
        "천포인트한판간다",
        "한비가비비빅을사오면서",
        "●▅▇█▇▆▅▄▇",
        "환자의용태에관한문제·0ƖςƐμटმ٢8୧진단0·1",
        "스텔라✩아르투아",
        "꽃들은천재지변이있더라도아랑곳하지않는다",
        "정원을갖게된후로시간이다르게흐른다",
        "내가더럽혀지더라도오직너에게흰것만을줄게",
        "너무뜨거워서다른사람이부담스러워하지도않고너무차가워서다른사람이상처받지도않는온도는따뜻함",
        "여름과함께떠나보낸너의그뒷모습은행복한꿈이었다말할테니"
    ]
    result = random.choice(sample)
    return result

def shufflePlayer(player, i):
    result = []
    result.append(player[i])
    print(result)
    player.pop(i)
    random.shuffle(player)
    result = result + player
    print(result)
    return result


TEAM_ALIASES = {
    "1팀": 1,
    "팀1": 1,
    "team1": 1,
    "team 1": 1,
    "2팀": 2,
    "팀2": 2,
    "team2": 2,
    "team 2": 2,
}
TEAM_RANK_POINTS = (10, 7, 5, 3, 2, 1, 0, -1)


def _parse_team_selection(value):
    """Split ``user reference + team`` while allowing a team-only owner choice."""
    text = str(value or "").strip()
    lowered = text.casefold()
    if lowered in TEAM_ALIASES:
        return None, TEAM_ALIASES[lowered]

    for alias, team_number in sorted(
        TEAM_ALIASES.items(), key=lambda item: len(item[0]), reverse=True
    ):
        suffix = f" {alias}"
        if lowered.endswith(suffix):
            return text[:-len(suffix)].strip(), team_number
    return text, None


def _balance_wordchain_teams(players):
    """Keep manual choices and randomly fill every remaining team slot."""
    team_size = len(players) // 2
    team1 = [player for player in players if player.get("team") == 1]
    team2 = [player for player in players if player.get("team") == 2]
    undecided = [player for player in players if player.get("team") not in (1, 2)]

    if len(team1) > team_size or len(team2) > team_size:
        raise ValueError("한 팀에 선택된 참가자가 너무 많습니다.")

    random.shuffle(undecided)
    while len(team1) < team_size:
        player = undecided.pop()
        player["team"] = 1
        team1.append(player)
    while len(team2) < team_size:
        player = undecided.pop()
        player["team"] = 2
        team2.append(player)

    return team1, team2


def _team_round_order(team1, team2):
    """Shuffle within teams, then alternate Team 1 and Team 2 players."""
    first = list(team1)
    second = list(team2)
    random.shuffle(first)
    random.shuffle(second)
    return [player for pair in zip(first, second) for player in pair]


def _team_roster_text(players, team_number=None):
    lines = []
    for player in players:
        if team_number is not None and player.get("team") != team_number:
            continue
        lv = etc.level(q.readXpById(player["id"]))
        team_label = f"{player['team']}팀" if player.get("team") else "자동 배정"
        lines.append(
            f"`{team_label}` {player_badge[player['color']]} "
            f"{scoreFont(player['score'], 4, player['color'])} "
            f"{lifeUI(player['life'], 3)} {etc.lvicon(lv)}"
            f"{q.readTagById(player['id'])}"
        )
    return "\n".join(lines)


def _apply_wordchain_achievements(player, rank, chain):
    user_id = player["id"]
    if l.wcReadById(user_id, "win") >= 84 and q.readStorageById(user_id, 81) == 0:
        q.storageModifyById(user_id, 81, 1)
    if chain > 421 and q.readStorageById(user_id, 83) == 0:
        q.storageModifyById(user_id, 83, 1)
    if l.wcReadById(user_id, "regist") >= 1446 and q.readStorageById(user_id, 84) == 0:
        q.storageModifyById(user_id, 84, 1)
    if player["score"] >= 1000 and q.readStorageById(user_id, 150) == 0:
        q.storageModifyById(user_id, 150, 1)
    if rank == 0 and player["life"] == 3 and q.readStorageById(user_id, 148) == 0:
        q.storageModifyById(user_id, 148, 1)
    elif rank == 0 and player["life"] == 0 and q.readStorageById(user_id, 149) == 0:
        q.storageModifyById(user_id, 149, 1)


async def _run_team_wordchain(client, ctx, players, option):
    """Run a team game using the same word and penalty rules as solo mode."""
    team1 = [player for player in players if player["team"] == 1]
    team2 = [player for player in players if player["team"] == 2]
    chain = 1
    history = []
    sample = sampleText()
    start = random.choice(sample)
    bonus = wd.random_korean()
    length = random.randint(2, 5)
    round_number = 0
    eliminated = None

    colors = list(range(1, 9))
    random.shuffle(colors)
    for player, color_number in zip(players, colors):
        player["color"] = color_number

    await ctx.send(
        f"**잠시 후 팀전이 시작됩니다!**\n"
        f"종목: 끝말잇기 {option}\n{_team_roster_text(players)}"
    )
    await asyncio.sleep(5)
    start_time = datetime.datetime.now().timestamp()

    while eliminated is None:
        round_number += 1
        turn_order = _team_round_order(team1, team2)
        await ctx.send(f"## {round_number} 라운드\n팀별 플레이 순서를 다시 섞었습니다.")

        for turn_index, current in enumerate(turn_order):
            uid = current["id"]
            ulv = etc.level(q.readXpById(uid))
            turn_failed = False

            if isOneKill(start):
                penalty = int(current["score"] * 0.42)
                current["score"] -= penalty
                current["life"] -= 1
                start = random.choice(sample)
                await ctx.send(
                    f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -{penalty}점** '
                    "한방 단어 공격을 받았습니다..."
                )
                turn_failed = True

            while current["life"] > 0 and not turn_failed:
                start_alter = replace_sound_char(start) or ""
                start_label = f"{start}({start_alter})" if start_alter else start
                rule_label = (
                    f"`보너스 글자` {bonus}"
                    if option == "일반"
                    else f"`제한 길이` {length}글자"
                )
                await ctx.send(
                    f":chains:{scoreFont(chain, 3, 0)} | `Team {current['team']}` "
                    f"{player_badge[current['color']]}{etc.lvicon(ulv)}"
                    f"{q.readTagById(uid)} | {scoreFont(current['score'], 4, current['color'])} "
                    f"| {lifeUI(current['life'], 3)} <@{uid}>\n"
                    f"## {start_label}\n{rule_label}\n"
                    "(으)로 시작하는 단어를 입력하세요! ('q' 입력시 포기)"
                )

                if uid == 691455977270149171:
                    if sample == "●▅▇█▇▆▅▄▇":
                        start = random.choice("가나다라마바사아자차카타파하")
                        start_alter = replace_sound_char(start) or ""
                    dice = random.randint(1, 2) if detectZwong(turn_index, turn_order) else random.randint(1, 7)
                    killer_length = length if option == "쿵쿵따" else 0
                    result = []
                    if dice == 1:
                        result = searchKiller(start, killer_length)
                        if start_alter:
                            result += searchKiller(start_alter, killer_length)
                    if result:
                        input_message = await ctx.send(random.choice(result))
                    else:
                        api_result = await asyncio.to_thread(
                            ks.startWord,
                            start,
                            history,
                            fixed_length=killer_length,
                        )
                        input_message = await ctx.send(api_result[0] if api_result else "q")
                else:
                    def message_check(message):
                        return message.author.id == uid and message.channel == ctx.channel

                    input_message = await client.wait_for("message", check=message_check)

                word = input_message.content.strip()
                if word == "q":
                    penalty = int(current["score"] * 0.33)
                    current["score"] -= penalty
                    current["life"] -= 1
                    start = random.choice(sample)
                    await ctx.send(
                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -{penalty}점** '
                        "방어에 실패하였습니다..."
                    )
                    turn_failed = True
                    continue

                valid_start = bool(word) and (
                    word[0] == start
                    or word[0] == start_alter
                    or sample == "●▅▇█▇▆▅▄▇"
                )
                if not valid_start:
                    current["life"] -= 1
                    current["score"] -= 30
                    await ctx.send(
                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** '
                        f"**`{start}`**(으)로 시작하는 단어를 입력해 주세요..."
                    )
                    continue
                if word in history:
                    current["life"] -= 1
                    current["score"] -= 50
                    await ctx.send(
                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -50점** '
                        "이미 사용한 단어입니다..."
                    )
                    turn_failed = True
                    continue
                if option == "일반" and len(word) < 2:
                    current["life"] -= 1
                    current["score"] -= 30
                    await ctx.send(
                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** '
                        "적어도 2글자 이상이어야 합니다..."
                    )
                    turn_failed = True
                    continue
                if option == "쿵쿵따" and len(word) != length:
                    current["life"] -= 1
                    current["score"] -= 30
                    await ctx.send(
                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** '
                        f"{length}글자 단어만 가능합니다..."
                    )
                    turn_failed = True
                    continue

                result = wd.readInGame(word)
                if result is None:
                    result = await asyncio.to_thread(ks.searchWord, word)
                if result is None:
                    current["life"] -= 1
                    current["score"] -= 30
                    await ctx.send(
                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** '
                        "없는 단어입니다..."
                    )
                    turn_failed = True
                    continue

                start = word[-1]
                history.append(word)
                gain = count_break_korean(word)
                if option == "일반" and bonus in word:
                    gain += 2 ** word.count(bonus)
                    bonus = wd.random_korean()
                current["score"] += scoreBoost(gain, current["life"])
                wd.newWordById(uid, str(result[0]), str(result[1]), str(result[2]))
                index = wd.findID(word)
                embed = discord.Embed(
                    title=f"{result[0]} `id: {index}`",
                    description=f"[{result[1]}] {result[2]}",
                    color=color[current["color"]],
                )
                embed.add_field(
                    name="**1팀 점수**",
                    value=_team_roster_text(players, 1),
                    inline=False,
                )
                embed.add_field(
                    name="**2팀 점수**",
                    value=_team_roster_text(players, 2),
                    inline=False,
                )
                embed.set_footer(text=f"{q.readTagById(uid)} | CHAIN: {chain}")
                await ctx.send(embed=embed)
                chain += 1
                break

            if current["life"] <= 0:
                eliminated = current
                await ctx.send(
                    f'`(⩌ʌ ⩌;)` **{q.readTagById(uid)}** 님이 탈락하여 게임이 종료됩니다.'
                )
                break

            if turn_failed:
                bonus = wd.random_korean()
                if option == "쿵쿵따":
                    length = random.randint(2, 5)

    elapsed = datetime.datetime.now().timestamp() - start_time
    elapsed_cs = int(elapsed * 100)
    for player in players:
        if player["life"] == 3:
            player["score"] = int(player["score"] * 2.4)
        player["score"] = max(0, player["score"])

    # Discrete rank points require a tie-break; shuffle first so recruitment
    # order does not consistently benefit players with an equal score.
    ranking = list(players)
    random.shuffle(ranking)
    ranking.sort(key=lambda player: -player["score"])
    team_rank_points = {1: 0, 2: 0}
    for rank, player in enumerate(ranking):
        player["rank_point"] = TEAM_RANK_POINTS[rank]
        team_rank_points[player["team"]] += player["rank_point"]

    team_scores = {
        1: sum(player["score"] for player in players if player["team"] == 1),
        2: sum(player["score"] for player in players if player["team"] == 2),
    }
    if team_rank_points[1] == team_rank_points[2]:
        winner_team = None
        result_title = "무승부"
    else:
        winner_team = max(team_rank_points, key=team_rank_points.get)
        result_title = f"{winner_team}팀 승리"

    embed = discord.Embed(
        title=f"팀전 결과 - {result_title}",
        description=(
            f"CHAIN: {chain - 1}\n"
            f"TIME: {elapsed_cs // 6000}분 {(elapsed_cs % 6000) // 100:02d}초 {elapsed_cs % 100:02d}\n\n"
            f"1팀: 승점 {team_rank_points[1]}점 | 합산 {team_scores[1]}점\n"
            f"2팀: 승점 {team_rank_points[2]}점 | 합산 {team_scores[2]}점"
        ),
        color=0xBCE29E,
    )

    for rank, player in enumerate(ranking):
        team_score = team_scores[player["team"]]
        reward_multiplier = 1.3 if winner_team == player["team"] else 1.0
        xp_gain = int((team_score * 1.8 + chain * 5) * reward_multiplier)
        money_gain = int((team_score * 1.2 + chain * 3) * reward_multiplier)
        q.xpAddById(player["id"], xp_gain)
        q.moneyAddById(player["id"], money_gain)
        is_winner = winner_team == player["team"]
        l.wcUpdateIndi(player["id"], player["score"], chain - 1, is_winner)
        _apply_wordchain_achievements(player, rank, chain)
        bonus_label = " | 승리 보너스 130%" if is_winner else ""
        embed.add_field(
            name=(
                f"{rank + 1}위 · {player['team']}팀 · "
                f"{q.readTagById(player['id'])}"
            ),
            value=(
                f"개인 {player['score']}점 | 승점 {player['rank_point']:+d}점 | "
                f"+{xp_gain}XP, +${money_gain:,}{bonus_label}"
            ),
            inline=False,
        )

    embed.set_footer(text="Discord Bot by Dizzt")
    await ctx.send("## 게임 끝", embed=embed)


# 단어목록 뷰어 UI

class PaginationList(discord.ui.View):
    current_page: int = 1
    sep: int = 20
    user = None

    async def send(self, ctx):
        self.message = await ctx.send(
            ":green_circle: ** 단어 검색이 완료되었습니다!",
            view=self)
        if self.current_page == 1:
            await self.update_message(self.data[:self.sep], self.user)
        elif self.current_page == int((len(self.data) - 1) / self.sep) + 1:
            await self.update_message(
                self.data[self.current_page * self.sep -
                          self.sep:len(self.data)], self.user)
        else:
            await self.update_message(
                self.data[(self.current_page - 1) *
                          self.sep:self.current_page * self.sep], self.user)

    def create_embed(self, data, user):
        embed = discord.Embed(
            title=f"**은비사전에 등록된 단어 목록**",
            description=f"총 {len(self.data)}개의 단어가 등록되었습니다.",
            color=0xFFFF72)
        
        result = ""

        for item in data:
            word = item[1]
            if len(word) > 12:
                word = item[1][:9] + "..."

            result += f"`{item[0]}` • **{word}** • {item[2]}\n"

        embed.add_field(
            name="",
            value=result[:-1],
            inline=False)

        embed.set_footer(
            text=
            f"Page : {self.current_page} / {int((len(self.data)-1) / self.sep) + 1}",
            icon_url="")

        return embed

    async def update_message(self, data, user):
        self.update_buttons()
        await self.message.edit(embed=self.create_embed(data, user), view=self)

    def update_buttons(self):
        if self.current_page == 1:
            self.first_page_button.disabled = True
            self.prev_button.disabled = True
            self.first_page_button.style = discord.ButtonStyle.gray
            self.prev_button.style = discord.ButtonStyle.gray
        else:
            self.first_page_button.disabled = False
            self.prev_button.disabled = False
            self.first_page_button.style = discord.ButtonStyle.green
            self.prev_button.style = discord.ButtonStyle.primary

        if self.current_page == int((len(self.data) - 1) / self.sep) + 1:
            self.next_button.disabled = True
            self.last_page_button.disabled = True
            self.last_page_button.style = discord.ButtonStyle.gray
            self.next_button.style = discord.ButtonStyle.gray
        else:
            self.next_button.disabled = False
            self.last_page_button.disabled = False
            self.last_page_button.style = discord.ButtonStyle.green
            self.next_button.style = discord.ButtonStyle.primary

    def get_current_page_data(self):
        until_item = self.current_page * self.sep
        from_item = until_item - self.sep
        if self.current_page == 1:
            from_item = 0
            until_item = self.sep
        if self.current_page == int((len(self.data) - 1) / self.sep) + 1:
            from_item = self.current_page * self.sep - self.sep
            until_item = len(self.data)
        return self.data[from_item:until_item]

    #맨 앞 페이지로 이동
    @discord.ui.button(label="|<", style=discord.ButtonStyle.green)
    async def first_page_button(self, interaction: discord.Interaction,
                                button: discord.ui.Button):
        if interaction.user == self.user:
            await interaction.response.defer()
            self.current_page = 1

            await self.update_message(self.get_current_page_data(), self.user)

    #앞 뒷 페이지로 이동
    @discord.ui.button(label="<", style=discord.ButtonStyle.primary)
    async def prev_button(self, interaction: discord.Interaction,
                          button: discord.ui.Button):
        if interaction.user == self.user:
            await interaction.response.defer()
            self.current_page -= 1
            await self.update_message(self.get_current_page_data(), self.user)

    #뒷 페이지로 이동
    @discord.ui.button(label=">", style=discord.ButtonStyle.primary)
    async def next_button(self, interaction: discord.Interaction,
                          button: discord.ui.Button):
        if interaction.user == self.user:
            await interaction.response.defer()
            self.current_page += 1
            await self.update_message(self.get_current_page_data(), self.user)

    #맨 뒷 페이지로 이동
    @discord.ui.button(label=">|", style=discord.ButtonStyle.green)
    async def last_page_button(self, interaction: discord.Interaction,
                               button: discord.ui.Button):
        if interaction.user == self.user:
            await interaction.response.defer()
            self.current_page = int((len(self.data) - 1) / self.sep) + 1
            await self.update_message(self.get_current_page_data(), self.user)


class TestCommands(commands.Cog):

    def __init__(self, client: commands.Bot):
        self.client = client

    # Wordcount Test [ID: 45]
    @commands.hybrid_command(name='점수', description="단어 점수를 계산합니다.")
    async def word_count_test(self, ctx, word: str = ''):
        if word != '':
            await ctx.reply(f'{word} => {count_break_korean(word)}')
        else:
            await ctx.reply('`(⩌ʌ ⩌;)` 비장상적인 테스트 케이스')

    @commands.hybrid_command(name='wcsupdate', description="단어 점수를 업데이트.")
    async def word_count_update(self, ctx, word: str = ''):
        wd.scoreUpdateAll()
        await ctx.reply('`(⩌ʌ ⩌;)` 완료')

    # Functions [ID: 33]
    @commands.hybrid_command(name='사전', description="우리말샘(국립국어원)에 실린 단어 뜻을 검색합니다!")
    async def word_search(self, ctx, *, word: str = ''):
        if word != '':
            result = await asyncio.to_thread(ks.searchWord, word)
            if result is not None:
                wd.newWord(ctx.author, str(result[0]), str(result[1]),
                           str(result[2]))

                temp = wd.readAll(word)
                embed = discord.Embed(title=temp[1],
                                        description=f'[{temp[2]}] {temp[3]}',
                                        color=0xBCE29E)
                embed.set_footer(
                    text=
                    f'색인번호: {temp[0]}\n등록일: {temp[5]}\n마지막 수정: {q.readTagById(temp[4])} ({temp[6]})'
                )
                await ctx.reply(':green_circle: 우리말샘(국립국어원) 검색 결과입니다.',
                                embed=embed)

            else:
                await ctx.reply('`(⩌Δ ⩌ ;)` 등록되지 않은 단어입니다...')

    @commands.hybrid_command(name='dict', description="단어 뜻을 검색합니다!")
    async def en_search(self, ctx, word: str = ''):
        if word != '':
            result = await asyncio.to_thread(ks.searchEn, word)
            if result is not None:
                await ctx.reply(f'## {result[0]}\n[{result[1]}] {result[2]}')
            else:
                await ctx.reply('`(⩌ʌ ⩌;)` 등록되지 않은 단어입니다...')

    # Functions [ID: 34]
    @commands.hybrid_command(name='은비사전', description="학습 된 단어 뜻을 검색합니다!")
    async def word_search_db(self,
                             ctx,
                             option: str = "검색",
                             *,
                             word: str = "*"):
        
        if option == "도움말":
            embed = discord.Embed(
                    title=f'은비사전 이용 가이드',
                    description=
                    '기본 형태는 `;은비사전 <옵션> <키워드/페이지>` 이고\n기본값은 `;은비사전 검색 *` 입니다.',
                    color=0x8ECDDD)

            embed.add_field(
                name=f"**도움말**",
                value="도움말을 열람합니다.\n따로 사용 가능한 키워드는 없습니다.",
                inline=False)
            
            embed.add_field(
                name=f"**검색**",
                value="단어를 검색합니다. 키워드 서식에 따라 조건 검색도 가능합니다.\n`<키워드>` 키워드가 일치하는 단어의 정보를 불러옵니다.\n`<글자>-` 해당 글자로 시작하는 단어들을 랜덤으로 골라 목록으로 보여줍니다.\n`-<글자>` 해당 글자로 끝나는 단어들을 랜덤으로 골라 목록으로 보여줍니다.\n`~<품사>` 해당 품사에 해당하는 단어들을 랜덤으로 골라 목록으로 보여줍니다.\n`id:<색인번호>` 해당 색인번호를 가진 단어의 정보를 불러옵니다.\n`%<패턴>` 패턴을 만족하는 단어의 정보를 불러옵니다.\n`*` 랜덤으로 10개의 단어를 골라 목록으로 보여줍니다.",
                inline=False)
            
            embed.add_field(
                name=f"**등록**",
                value="사용자 지정 단어를 등록합니다.\n기본 형식은 `;은비사전 등록 <단어>/<품사/주제>/<뜻>`입니다.\n`;은비사전 등록`을 통해 등록 요령을 볼 수 있습니다.",
                inline=False)
            
            embed.add_field(
                name=f"**수정**",
                value="사용자 지정 단어의 데이터를 수정합니다.\n기본 형식은 `;은비사전 수정 <색인번호>/<단어>/<품사/주제>/<뜻>`입니다.\n`;은비사전 수정`을 통해 정보 수정 요령을 볼 수 있습니다.",
                inline=False)
            
            embed.add_field(
                name=f"**품사변경**",
                value="품사 데이터를 일괄적으로 변경합니다.\n기본 형식은 `;은비사전 품사변경 <변경전>/<변경후>`입니다.\n`;은비사전 품사변경`을 통해 정보 수정 요령을 볼 수 있습니다.",
                inline=False)
            
            embed.add_field(
                name=f"**목록**",
                value="검색 조건에 해당하는 전체 단어의 목록을 불러옵니다.\n기본 형식은 `;은비사전 목록 <검색조건>`입니다.\n`검색조건`은 `<페이지>`, `<품사>/<페이지>`, `점수/<페이지>`, `품사/<페이지>, `%<패턴>/<페이지>` 가 있습니다.",
                inline=False)

            embed.set_footer(text='Discord Bot by Dizzt')

            await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)


        elif option == "검색":
            if word[-1] == "-" and len(word) == 2:
                result = wd.readAllByStart(word[0])
                if len(result) > 10:
                    temp = random.sample(result, 10)
                else:
                    temp = result

                embed = discord.Embed(
                    title=f'{word[0]}(으)로 시작하는 단어',
                    description=
                    f'총 {len(result)}개 중 {len(temp)}개를 무작위로 들고왔습니다!',
                    color=0x8ECDDD)

                for info in temp:
                    embed.add_field(
                        name=f"**{info[1]}**",
                        value=
                        f"ID: {info[0]} | 점수: {count_break_korean(info[1])}",
                        inline=False)

                embed.set_footer(text='Discord Bot by Dizzt')
                await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)

            elif word[-1] == "!" and len(word) == 2:
                result = searchKiller(word[0], 0)
                print(result)
                if len(result) > 10:
                    temp = random.sample(result, 10)
                else:
                    temp = result

                embed = discord.Embed(
                    title=f'{word[0]}(으)로 시작하는 공격단어',
                    description=
                    f'총 {len(result)}개 중 {len(temp)}개를 무작위로 들고왔습니다!',
                    color=0x8ECDDD)

                for word in temp:
                    info = wd.readAll(word)
                    embed.add_field(
                        name=f"**{info[1]}**",
                        value=
                        f"ID: {info[0]} | 점수: {count_break_korean(info[1])}",
                        inline=False)

                embed.set_footer(text='Discord Bot by Dizzt')
                await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)

            elif word[0] == "-" and len(word) == 2:
                result = wd.readAllByEnd(word[-1])
                if len(result) > 10:
                    temp = random.sample(result, 10)
                else:
                    temp = result

                embed = discord.Embed(
                    title=f'{word[-1]}(으)로 끝나는 단어',
                    description=
                    f'총 {len(result)}개 중 {len(temp)}개를 무작위로 들고왔습니다!',
                    color=0x8ECDDD)

                for info in temp:
                    embed.add_field(
                        name=f"**{info[1]}**",
                        value=
                        f"ID: {info[0]} | 점수: {count_break_korean(info[1])}",
                        inline=False)

                embed.set_footer(text='Discord Bot by Dizzt')
                await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)

            elif word[0] == "~":
                result = wd.readAllByPOS(word[1:])
                if len(result) > 10:
                    temp = random.sample(result, 10)
                else:
                    temp = result

                embed = discord.Embed(
                    title=f'품사가 {word[1:]}인 단어',
                    description=
                    f'총 {len(result)}개 중 {len(temp)}개를 무작위로 들고왔습니다!',
                    color=0x8ECDDD)

                for info in temp:
                    embed.add_field(
                        name=f"**{info[1]}**",
                        value=
                        f"ID: {info[0]} | 점수: {count_break_korean(info[1])}",
                        inline=False)

                embed.set_footer(text='Discord Bot by Dizzt')
                await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)
            
            elif word[0] == "%":
                result = wd.readAllPattern(word[1:])
                if len(result) > 10:
                    temp = random.sample(result, 10)
                else:
                    temp = result

                embed = discord.Embed(
                    title=f'패턴 "{word[1:]}"을 만족하는 단어',
                    description=
                    f'총 {len(result)}개 중 {len(temp)}개를 무작위로 들고왔습니다!',
                    color=0x8ECDDD)

                for info in temp:
                    embed.add_field(
                        name=f"**{info[1]}**",
                        value=
                        f"ID: {info[0]} | 점수: {count_break_korean(info[1])}",
                        inline=False)

                embed.set_footer(text='Discord Bot by Dizzt')
                await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)

            elif word[0:3] == "id:":
                val = int(word[3:])
                result = wd.readAllById(val)
                if result is not None:
                    point = count_break_korean(result[1])
                    embed = discord.Embed(
                        title=result[1],
                        description=f'[{result[2]}] {result[3]}\n\n획득가능 점수: 3목숨 (+0%) **{point}점** | 2목숨 (+16%) **{int(1.16*point)}점** | 1목숨 (+39%) **{int(1.39*point)}점**',
                        color=0x8ECDDD)
                    embed.set_footer(
                        text=
                        f'색인번호: {result[0]}\n등록일: {result[5]}\n마지막 수정: {q.readTagById(result[4])} ({result[6]})'
                    )
                    await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!',
                                    embed=embed)
                else:
                    await ctx.reply(
                        '## `(⩌Δ ⩌ ;)` 없는 단어 입니다...\n 혹시 단어를 새로 등록해 보는 것은 어떨까요?'
                    )

            elif word == "*":
                result = wd.readAllRandom()
                temp = random.sample(result, 10)
                print(temp)
                embed = discord.Embed(
                    title='랜덤 단어 생성',
                    description=
                    f'총 {len(result)}개 중 {len(temp)}개를 무작위로 들고왔습니다!',
                    color=0x8ECDDD)

                for info in temp:
                    embed.add_field(
                        name=f"**{info[1]}**",
                        value=
                        f"ID: {info[0]} | 점수: {count_break_korean(info[1])}",
                        inline=False)

                embed.set_footer(text='Discord Bot by Dizzt')
                await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!', embed=embed)

            elif word != "":
                result = wd.readAll(word)
                if result is not None:
                    point = count_break_korean(result[1])
                    embed = discord.Embed(
                        title=result[1],
                        description=f'[{result[2]}] {result[3]}\n\n획득가능 점수: 3목숨 (+0%) **{point}점** | 2목숨 (+16%) **{int(1.16*point)}점** | 1목숨 (+39%) **{int(1.39*point)}점**',
                        color=0x8ECDDD)
                    embed.set_footer(
                        text=
                        f'색인번호: {result[0]}\n등록일: {result[5]}\n마지막 수정: {q.readTagById(result[4])} ({result[6]})'
                    )
                    await ctx.reply(':green_circle: 단어 검색이 완료되었습니다!',
                                    embed=embed)
                else:
                    await ctx.reply(
                        '## `(⩌Δ ⩌ ;)` 없는 단어 입니다...\n 혹시 단어를 새로 등록해 보는 것은 어떨까요?'
                    )

        elif option == "등록":
            if word == "도움말":
                await ctx.reply('`은비사전 등록 <단어>/<품사>/<뜻>` 구문으로 단어 등록이 가능합니다!')
            elif word != "":
                try:
                    result = word.split("/")
                    text = re.sub('[^A-Za-z0-9가-힣ㄱ-ㆎ]', '', result[0])
                    wd.newWord(ctx.author, text, result[1], result[2])

                    temp = wd.readAll(result[0])
                    embed = discord.Embed(title=temp[1],
                                          description=f'[{temp[2]}] {temp[3]}',
                                          color=0xBCE29E)
                    embed.set_footer(
                        text=
                        f'색인번호: {temp[0]}\n등록일: {temp[5]}\n마지막 수정: {q.readTagById(temp[4])} ({temp[6]})'
                    )
                    await ctx.reply(':green_circle: 단어 등록이 완료되었습니다!',
                                    embed=embed)
                except:
                    await ctx.reply(
                        '## `(⩌Δ ⩌ ;)` 단어 등록에 실패 하였습니다...\n* `은비사전 검색 <단어>`로 이미 등록된 단어인지 확인해 주세요.\n* `은비사전 등록 <단어>/<품사>/<뜻>` 구문이 정확한지 확인해 주세요.'
                    )

        elif option == "동의어":
            if word == "도움말":
                await ctx.reply('`은비사전 동의어 <색인번호>/<동의어로 등록 될 단어>` 구문으로 단어 등록이 가능합니다!')
            elif word != "":
                try:
                    result = word.split("/")
                    text = re.sub('[^A-Za-z0-9가-힣ㄱ-ㆎ]', '', result[1])
                    desc = wd.readAllById(int(result[0]))
                    wd.newWord(ctx.author, text, desc[2], desc[3])

                    temp = wd.readAll(text)
                    embed = discord.Embed(title=temp[1],
                                          description=f'[{temp[2]}] {temp[3]}',
                                          color=0xBCE29E)
                    embed.set_footer(
                        text=
                        f'색인번호: {temp[0]}\n등록일: {temp[5]}\n마지막 수정: {q.readTagById(temp[4])} ({temp[6]})'
                    )
                    await ctx.reply(':green_circle: 단어 등록이 완료되었습니다!',
                                    embed=embed)
                except:
                    await ctx.reply(
                        '## `(⩌Δ ⩌ ;)` 단어 등록에 실패 하였습니다...\n* `은비사전 검색 <단어>`로 이미 등록된 단어인지 확인해 주세요.\n* `은비사전 등록 <단어>/<품사>/<뜻>` 구문이 정확한지 확인해 주세요.'
                    )

        elif option == "수정":
            if word == "도움말":
                await ctx.reply(
                    '`은비사전 수정 <색인번호>/<단어>/<품사>/<뜻>` 구문으로 단어 수정이 가능합니다.\n`수정 후` 인자들은 비워두면 수정이 되지않고 전 데이터를 유지하게 됩니다.\n`뜻` 항목 맨앞에 `+`를 입력하여 뒤에 이어쓰기를 할 수 있습니다.'
                )
            elif word != "":
                try:
                    result = word.split("/")

                    if result[1] != "":
                        text = re.sub('[^A-Za-z0-9가-힣ㄱ-ㆎ]', '', result[1])
                        wd.wordModify(ctx.author, int(result[0]), text)

                    if result[2] != "":
                        wd.plModify(ctx.author, int(result[0]), result[2])

                    if result[3] != "":
                        wd.meanModify(ctx.author, int(result[0]), result[3])

                    temp = wd.readAllById(int(result[0]))
                    embed = discord.Embed(title=temp[1],
                                          description=f'[{temp[2]}] {temp[3]}',
                                          color=0xFFCC70)
                    embed.set_footer(
                        text=
                        f'색인번호: {temp[0]}\n등록일: ({temp[5]})\n수정: {q.readTagById(temp[4])} ({temp[6]})'
                    )
                    await ctx.reply(':green_circle: 단어 수정이 완료되었습니다!',
                                    embed=embed)
                    q.xpAdd(ctx.author, 100)
                    q.moneyAdd(ctx.author, 30)
                except:
                    await ctx.reply(
                        '## `(⩌Δ ⩌ ;)` 단어 수정에 실패 하였습니다...\n* `은비사전 검색 <단어>`로 등록되지 않은 단어인지 확인해 주세요.\n* `<전단어>/<후단어>/<품사>/<뜻>` 구문이 정확한지 확인해 주세요.'
                    )

        elif option == "품사변경":

            if word == "도움말":
                await ctx.reply(
                    '`은비사전 품사변경 <변경전>/<변경후>` 구문으로 품사 변경이 가능합니다.\n`변경전` 인자들은 비워두면 품사가 지정되지 않은 단어들이 선택됩니다.'
                )
            elif word != "":
                temp = word.split('/')
                print(temp)
                if len(temp) == 2:
                    wd.categoryModify(ctx.author, temp[0], temp[1])
                    await ctx.reply(f':green_circle: 품사 수정({temp[0]} → {temp[1]})이 완료되었습니다!')
                else:
                    await ctx.reply(
                            '## `(⩌Δ ⩌ ;)` 단어 수정에 실패 하였습니다...\n* `은비사전 품사변경 <변경전>/<변경후>` 구문이 정확한지 확인해 주세요.'
                        )

        elif option == "목록":

            temp = word.split('/')

            if len(temp) == 2:
                try:
                    page = int(temp[1])
                except:
                    None
                if temp[0][0] == "%":
                    pagination_view = PaginationList(timeout=None)
                    pagination_view.data = wd.readAllPattern(temp[0][1:])
                    pagination_view.user = ctx.author
                    pagination_view.current_page = page
                    await pagination_view.send(ctx)
                elif temp[0] == "점수":
                    pagination_view = PaginationList(timeout=None)
                    pagination_view.data = wd.readAllScore()
                    pagination_view.user = ctx.author
                    pagination_view.current_page = page
                    await pagination_view.send(ctx)
                elif temp[0] == "품사":
                    pagination_view = PaginationList(timeout=None)
                    pagination_view.data = wd.readPOSCount()
                    pagination_view.user = ctx.author
                    pagination_view.current_page = page
                    await pagination_view.send(ctx)
                else:
                    pagination_view = PaginationList(timeout=None)
                    pagination_view.data = wd.readAllByPOSWithPOS(temp[0])
                    pagination_view.user = ctx.author
                    pagination_view.current_page = page
                    await pagination_view.send(ctx)

            else:
                page = 1
                try:
                    page = int(word)
                except:
                    None
                pagination_view = PaginationList(timeout=None)
                pagination_view.data = wd.readAllWithPOS()
                pagination_view.user = ctx.author
                pagination_view.current_page = page
                await pagination_view.send(ctx)
            

    # Functions [ID: 87]
    @commands.hybrid_command(name='일괄수정',
                             description="엑셀에 저장된 파일을 일괄적으로 업데이트 합니다!")
    async def word_edit_db(self, ctx, uid: int = None, file: str = ""):
        total = 0
        suc_count = 0
        fail_count = 0
        path = DATA_DIR / "word_enlist" / f"{file}.xlsx"
        result = []
        bool = True

        try:
            await ctx.send(f"[1/3] {file}.xlsx 파일을 찾고있습니다!")
            await ctx.send(f"[2/3] {file}.xlsx 파일을 읽고있습니다!")
            rows = await asyncio.to_thread(_read_workbook_rows, path)
            for row in rows:
                result.append(list(row[:3]))
            del rows
        except:
            await ctx.send(i18n.t(ctx.author, "common.file_error", file=f"{file}.xlsx"))
            bool = False

        if bool:
            total = len(result)
            await ctx.send(
                f"[3/3] {total}개의 파일을 찾았습니다. 사전에 데이터를 추가합니다! 데이터가 많으면 시간이 좀 오래 걸립니다!"
            )

            for data in result:
                try:
                    index = wd.readAll(data[0])[0]
                    if data[2] == "우리말샘":
                        try:
                            search_result = await asyncio.to_thread(
                                ks.searchWord, data[0]
                            )
                            mean = search_result[2]
                            wd.plModify(ctx.author, int(index), data[1])
                            wd.meanModify(ctx.author, int(index), mean)
                        except: #mean이 none 일 경우
                            mean = f"{data[1]} 관련 단어."
                            wd.plModify(ctx.author, int(index), data[1])
                            wd.meanModify(ctx.author, int(index), mean)
                    else:
                        wd.plModify(ctx.author, int(index), data[1])
                        wd.meanModify(ctx.author, int(index), data[2])

                    temp = wd.readAll(data[0])
                    embed = discord.Embed(title=temp[1],
                                          description=f'[{temp[2]}] {temp[3]}',
                                          color=0xBCE29E)
                    embed.set_footer(
                        text=
                        f'색인번호: {temp[0]}\n등록일: {temp[5]}\n마지막 수정: {q.readTagById(temp[4])} ({temp[6]})'
                    )
                    suc_count += 1
                    await ctx.send(
                        f'[{suc_count+fail_count}/{total}] 단어 수정이 완료되었습니다.',
                        embed=embed)
                except:
                    fail_count += 1
                    await ctx.send(
                        f'[{suc_count+fail_count}/{total}] 서식에 오류가 있는 단어 입니다.'
                    )

            await ctx.send(
                i18n.t(ctx.author, "common.batch_complete", total=total, success=suc_count, failure=fail_count)
            )

    # Functions [ID: 87]
    @commands.hybrid_command(name='일괄등록',
                             description="엑셀에 저장된 파일을 일괄적으로 업데이트 합니다!")
    async def word_enlist_db(self, ctx, uid: int = None, file: str = "", option:int = 0):
        total = 0
        suc_count = 0
        fail_count = 0
        path = DATA_DIR / "word_enlist" / f"{file}.xlsx"
        result = []
        bool = True

        try:
            await ctx.send(f"[1/3] {file}.xlsx 파일을 찾고있습니다!")
            await ctx.send(f"[2/3] {file}.xlsx 파일을 읽고있습니다!")
            rows = await asyncio.to_thread(_read_workbook_rows, path)
            for row in rows:
                result.append(list(row[:3]))
            del rows
        except:
            await ctx.send(i18n.t(ctx.author, "common.file_error", file=f"{file}.xlsx"))
            bool = False

        if bool:
            total = len(result)
            await ctx.send(
                f"[3/3] {total}개의 파일을 찾았습니다. 사전에 데이터를 추가합니다! 데이터가 많으면 시간이 좀 오래 걸립니다!"
            )

            for data in result:
                try:
                    if data[2] == "우리말샘":
                        try:
                            search_result = await asyncio.to_thread(
                                ks.searchWord, data[0]
                            )
                            mean = search_result[2]
                            wd.newWordById(uid, data[0], data[1], mean)
                        except: #mean이 none 일 경우
                            mean = f"{data[1]} 관련 단어."
                            wd.newWordById(uid, data[0], data[1], mean)
                    else:
                        wd.newWordById(uid, data[0], data[1], data[2])

                    temp = wd.readAll(data[0])
                    suc_count += 1

                    if option == 0 or data[1] != temp[2]:
                        embed = discord.Embed(title=temp[1],
                                            description=f'[{temp[2]}] {temp[3]}',
                                            color=0xBCE29E)
                        embed.set_footer(
                            text=
                            f'색인번호: {temp[0]}\n등록일: {temp[5]}\n마지막 수정: {q.readTagById(temp[4])} ({temp[6]})'
                        )
                        await ctx.send(
                            f'[{suc_count+fail_count}/{total}] 단어 등록이 완료되었습니다.',
                            embed=embed)
                except:
                    fail_count += 1
                    await ctx.send(
                        f'[{suc_count+fail_count}/{total}] 이미 등록된 단어이거나 서식에 오류가 있는 단어 입니다.'
                    )

            await ctx.send(
                i18n.t(ctx.author, "common.batch_complete", total=total, success=suc_count, failure=fail_count)
            )

    @commands.hybrid_command(name='조건등록',
                             description="엑셀에 저장된 파일을 일괄적으로 업데이트 합니다!")
    async def word_c_enlist_db(self, ctx, uid: int = None, file: str = ""):
        total = 0
        suc_count = 0
        fail_count = 0
        path = DATA_DIR / "word_enlist" / f"{file}.xlsx"
        result = []
        bool = True

        try:
            await ctx.send(f"[1/3] {file}.xlsx 파일을 찾고있습니다!")
            await ctx.send(f"[2/3] {file}.xlsx 파일을 읽고있습니다!")
            rows = await asyncio.to_thread(_read_workbook_rows, path)
            for row in rows:
                result.append(list(row[:3]))
            del rows
        except:
            await ctx.send(i18n.t(ctx.author, "common.file_error", file=f"{file}.xlsx"))
            bool = False

        if bool:
            total = len(result)
            await ctx.send(
                f"[3/3] {total}개의 파일을 찾았습니다. 사전에 데이터를 추가합니다! 데이터가 많으면 시간이 좀 오래 걸립니다!"
            )

            for data in result:
                try:
                    if data[2] == "우리말샘":
                        try:
                            search_result = await asyncio.to_thread(
                                ks.searchWord, data[0]
                            )
                            mean = search_result[2]
                            pos = search_result[1]
                            wd.newWordById(uid, data[0], pos, mean)
                        except: #mean이 none 일 경우
                            mean = f"{data[1]} 관련 단어."
                            wd.newWordById(uid, data[0], data[1], mean)
                    else:
                        wd.newWordById(uid, data[0], data[1], data[2])

                    temp = wd.readAll(data[0])
                    embed = discord.Embed(title=temp[1],
                                          description=f'[{temp[2]}] {temp[3]}',
                                          color=0xBCE29E)
                    embed.set_footer(
                        text=
                        f'색인번호: {temp[0]}\n등록일: {temp[5]}\n마지막 수정: {q.readTagById(temp[4])} ({temp[6]})'
                    )
                    suc_count += 1
                    await ctx.send(
                        f'[{suc_count+fail_count}/{total}] 단어 등록이 완료되었습니다.',
                        embed=embed)
                except:
                    fail_count += 1
                    await ctx.send(
                        f'[{suc_count+fail_count}/{total}] 이미 등록된 단어이거나 서식에 오류가 있는 단어 입니다.'
                    )

            await ctx.send(
                i18n.t(ctx.author, "common.batch_complete", total=total, success=suc_count, failure=fail_count)
            )

    @commands.hybrid_command(name='빠른등록',
                             description="엑셀에 저장된 파일을 일괄적으로 업데이트 합니다!")
    async def word_q_enlist_db(self, ctx, uid: int = None, file: str = ""):
        total = 0
        suc_count = 0
        fail_count = 0
        path = DATA_DIR / "word_enlist" / f"{file}.xlsx"
        result = []
        bool = True

        try:
            await ctx.send(f"[1/3] {file}.xlsx 파일을 찾고있습니다!")
            await ctx.send(f"[2/3] {file}.xlsx 파일을 읽고있습니다!")
            rows = await asyncio.to_thread(_read_workbook_rows, path)
            for row in rows:
                result.append(list(row[:3]))
            del rows
        except:
            await ctx.send(i18n.t(ctx.author, "common.file_error", file=f"{file}.xlsx"))
            bool = False

        if bool:
            total = len(result)
            await ctx.send(
                f"[3/3] {total}개의 파일을 찾았습니다. 사전에 데이터를 추가합니다! 데이터가 많으면 시간이 좀 오래 걸립니다!"
            )

            for data in result:
                try:
                    if data[2] == "우리말샘":
                        try:
                            search_result = await asyncio.to_thread(
                                ks.searchWord, data[0]
                            )
                            mean = search_result[2]
                            wd.newWordById(uid, data[0], data[1], mean)
                        except: #mean이 none 일 경우
                            mean = f"{data[1]} 관련 단어."
                            wd.newWordById(uid, data[0], data[1], mean)
                    else:
                        wd.newWordById(uid, data[0], data[1], data[2])

                    suc_count += 1
                except:
                    fail_count += 1
                print(f"[{suc_count+fail_count}/{total} | {(suc_count+fail_count)/total:.2%}] 작업 완료")

            await ctx.send(
                i18n.t(ctx.author, "common.batch_complete", total=total, success=suc_count, failure=fail_count)
            )

    @commands.hybrid_command(name='빠른조건등록',
                             description="엑셀에 저장된 파일을 일괄적으로 업데이트 합니다!")
    async def word_qc_enlist_db(self, ctx, uid: int = None, file: str = ""):
        total = 0
        suc_count = 0
        fail_count = 0
        path = DATA_DIR / "word_enlist" / f"{file}.xlsx"
        result = []
        bool = True

        try:
            await ctx.send(f"[1/3] {file}.xlsx 파일을 찾고있습니다!")
            await ctx.send(f"[2/3] {file}.xlsx 파일을 읽고있습니다!")
            rows = await asyncio.to_thread(_read_workbook_rows, path)
            for row in rows:
                result.append(list(row[:3]))
            del rows
        except:
            await ctx.send(i18n.t(ctx.author, "common.file_error", file=f"{file}.xlsx"))
            bool = False

        if bool:
            total = len(result)
            await ctx.send(
                f"[3/3] {total}개의 파일을 찾았습니다. 사전에 데이터를 추가합니다! 데이터가 많으면 시간이 좀 오래 걸립니다!"
            )

            for data in result:
                try:
                    if data[2] == "우리말샘":
                        try:
                            search_result = await asyncio.to_thread(
                                ks.searchWord, data[0]
                            )
                            mean = search_result[2]
                            pos = search_result[1]
                            wd.newWordById(uid, data[0], pos, mean)
                        except: #mean이 none 일 경우
                            mean = f"{data[1]} 관련 단어."
                            wd.newWordById(uid, data[0], data[1], mean)
                    else:
                        wd.newWordById(uid, data[0], data[1], data[2])

                    suc_count += 1
                except:
                    fail_count += 1
                print(f"[{suc_count+fail_count}/{total} | {(suc_count+fail_count)/total:.2%}] 작업 완료")

            await ctx.send(
                i18n.t(ctx.author, "common.batch_complete", total=total, success=suc_count, failure=fail_count)
            )


    # Functions [ID: 42]
    @commands.hybrid_command(name='끝말잇기', description="일반 끝말잇기를 시작합니다!")
    async def word_chain(self, ctx, option: str = '일반', team: str = '개인'):
        if team == '개인':
            if option in ['일반', '쿵쿵따']:
                gamestart = False
                not_include = []
                player = []
                owner, _ = await _prepare_wordchain_participant(ctx)
                player.append({"id": owner.id, "score": 0, "life": 3, "color": 0})
                not_include.append(owner.id)

                while True:
                    embed = discord.Embed(title='참가자 목록',
                                            description=f'인원: {len(player)}/8',
                                            color=0xBCE29E)
                    for i in range(len(player)):
                        lv = etc.level(q.readXpById(player[i]['id']))
                        sts = l.wcReadById(player[i]['id'], 'stats')
                        print(sts)
                        embed.add_field(
                            name=
                            f"{i+1}. {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}",
                            value=f"`전적` {sts[2]}전 {sts[3]}승 | {sts[0]}점 | {sts[1]}체인",
                            inline=False)
                    embed.set_footer(text='Discord Bot by Dizzt')
                    await ctx.reply(
                        f"## 끝말잇기 ({option}) - 인원 모집\n* `@username`을 이용하여 최대 8명 까지 초대가 가능합니다!\n* 초대가 완료되면 `시작`를 입력해 주세요!\n* 게임 생성을 취소하고 싶다면 `취소`를 입력해 주세요!\n* 게임이 시작되면 자동으로 순서가 바뀝니다!",
                        embed=embed)

                    def check(m):
                        return m.author == ctx.author and m.channel == ctx.channel

                    input_word = await self.client.wait_for("message", check=check)
                    check = input_word.content

                    if check == '시작':
                        if len(player) > 1:
                            gamestart = True
                            await ctx.send(":green_circle: 게임이 성공적으로 생성 되었습니다!")
                            break
                        else:
                            await ctx.send(
                                '`(⩌ʌ ⩌;)` 인원이 너무 적습니다... 적어도 2명 이상 있어야 합니다!')

                    elif check == '취소':
                        await ctx.send(":x: 게임 생성이 취소되었습니다.")
                        break

                    else:
                        try:
                            if len(player) >= 8:
                                await ctx.send(
                                    '`(⩌ʌ ⩌;)` 인원이 너무 많습니다... 최대 8명 까지 참가가 가능합니다!')                            
                            else:
                                invited, _ = await _prepare_wordchain_participant(ctx, check)
                                id = invited.id
                                if id in not_include:
                                    await ctx.send(
                                    '`(⩌ʌ ⩌;)` 이미 참가한 사람입니다...')
                                else:
                                    name = q.readTagById(id)
                                    player.append({"id": id, "score": 0, "life": 3, "color": 0})
                                    not_include.append(id)
                                    await ctx.send(
                                        f":green_circle: `{name}`가 참가자 목록에 추가되었습니다! ")
                        except (UserResolutionError, ValueError):
                            await ctx.send(
                                '`(⩌ʌ ⩌;)` 유효하지 않은 참가자 입니다... 다시 시도해 보세요...')

                if gamestart and option == '일반':
                    random.shuffle(player)
                    chain = 1
                    history = []
                    sample = sampleText()
                    start = random.choice(sample)
                    start_alter = ""
                    bonus = wd.random_korean()
                    end = False
                    shCheck = False

                    color_arr = [1, 2, 3, 4, 5 ,6, 7, 8]
                    random.shuffle(color_arr)

                    for i in range(len(player)):
                        player[i]['color'] = color_arr[i]

                    embed = discord.Embed(title='참가자 목록',
                                            description=f'인원: {len(player)}/6',
                                            color=0xBCE29E)
                    for i in range(len(player)):
                        lv = etc.level(q.readXpById(player[i]['id']))
                        embed.add_field(
                            name=
                            f"{player_badge[player[i]['color']]}{etc.lvicon(lv)}{q.readTagById(player[i]['id'])} (Lv. {lv})",
                            value=
                            f"**{scoreFont(player[i]['score'], 4, player[i]['color'])}점** | {lifeUI(player[i]['life'],3)}",
                            inline=False)
                    embed.set_footer(text='Discord Bot by Dizzt')
                    await ctx.send(
                        f"**잠시후 게임이 시작됩니다!**\n종목: 끝말잇기 (**`{sample}`** 중 한 글자가 랜덤으로 배치 됩니다.)"
                    )
                    await asyncio.sleep(5)
                    start_time = datetime.datetime.now().timestamp()

                    while True:
                        for i in range(len(player)):

                            uid = player[i]['id']
                            ulv = etc.level(q.readXpById(uid))

                            if player[i]['life'] > 0 and isOneKill(start):
                                s = int(player[i]['score'] * 0.42)
                                player[i]['score'] -= s
                                player[i]['life'] -= 1
                                start = random.choice(sample)
                                await ctx.send(
                                    f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -{s}점** 한방 단어 공격을 받았습니다...'
                                )
                                player = shufflePlayer(player, i)
                                bonus = wd.random_korean()
                                break

                            while player[i]['life'] > 0:

                                if replace_sound_char(start) is not None:
                                    start_alter = replace_sound_char(start)
                                    await ctx.send(
                                        f":chains:{scoreFont(chain, 3, 0)} | {player_badge[player[i]['color']]}{etc.lvicon(ulv)}{q.readTagById(uid)} | {scoreFont(player[i]['score'], 4, player[i]['color'])} | {lifeUI(player[i]['life'],3)} <@{uid}>\n## {start}({start_alter})\n`보너스 글자` {bonus}\n(으)로 시작하는 단어를 입력하세요! ('q' 입력시 포기)"
                                    )
                                else:
                                    start_alter = ""
                                    await ctx.send(
                                        f":chains:{scoreFont(chain, 3, 0)} | {player_badge[player[i]['color']]}{etc.lvicon(ulv)}{q.readTagById(uid)} | {scoreFont(player[i]['score'], 4, player[i]['color'])} | {lifeUI(player[i]['life'],3)} <@{uid}>\n## {start}\n`보너스 글자` {bonus}\n(으)로 시작하는 단어를 입력하세요! ('q' 입력시 포기)"
                                    )

                                if player[i]['id'] == 691455977270149171:
                                    if sample == '●▅▇█▇▆▅▄▇':
                                        alterlist = "가나다라마사바아자파카타파하"
                                        pick = random.randint(1, len(alterlist))
                                        start = alterlist[pick]

                                    dice = 0
                                    if detectZwong(i, player):
                                        dice = random.randint(1, 2)
                                    else:
                                        dice = random.randint(1, 7)
                                    start_alter = replace_sound_char(start)
                                    if dice == 1:
                                        result = searchKiller(start, 0)
                                        if start_alter is not None:
                                            result += searchKiller(start_alter, 0)
                                        if len(result) != 0:
                                            input_word = await ctx.send(random.choice(result))
                                        else:
                                            result = await asyncio.to_thread(
                                                ks.startWord, start, history
                                            )
                                            if result is not None:
                                                input_word = await ctx.send(result[0])
                                            else:
                                                input_word = await ctx.send("q")
                                    else:
                                        result = await asyncio.to_thread(
                                            ks.startWord, start, history
                                        )
                                        if result is not None:
                                            input_word = await ctx.send(result[0])
                                        else:
                                            input_word = await ctx.send("q")

                                else:

                                    def check(m):
                                        return m.author.id == uid and m.channel == ctx.channel

                                    input_word = await self.client.wait_for(
                                        "message", check=check)

                                check = input_word.content

                                if check == 'q':
                                    s = int(player[i]['score'] * 0.33)
                                    player[i]['score'] -= s
                                    player[i]['life'] -= 1
                                    start = random.choice(sample)
                                    await ctx.send(
                                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -{s}점** 방어에 실패하였습니다...'
                                    )
                                    shCheck = True
                                    break

                                else:
                                    if (
                                        check[0] == start
                                        or check[0] == start_alter
                                        or sample == '●▅▇█▇▆▅▄▇'
                                    ) and len(check) > 1 and check not in history:

                                        result = wd.readInGame(check)

                                        if result is None:
                                            result = await asyncio.to_thread(
                                                ks.searchWord, check
                                            )

                                        if result is not None:
                                            start = check[-1]
                                            history.append(check)
                                            gain = count_break_korean(check)

                                            if bonus in check:
                                                gain += 2**check.count(bonus)
                                                bonus = wd.random_korean()

                                            player[i]['score'] += scoreBoost(
                                                gain,
                                                player[i]['life'])
                                            wd.newWordById(player[i]['id'],
                                                            str(result[0]),
                                                            str(result[1]),
                                                            str(result[2]))
                                            index = wd.findID(check)
                                            name = q.readTagById(player[i]['id'])
                                            embed = discord.Embed(
                                                title=f'{result[0]} `id: {index}`',
                                                description=
                                                f'[{result[1]}] {result[2]}',
                                                color=color[player[i]['color']])
                                            
                                            if len(player) < 5:
                                                text = ""
                                                for i in range(len(player)):
                                                    lv = etc.level(
                                                        q.readXpById(player[i]['id']))
                                                    if i == 0:
                                                        text += f"{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                    else:
                                                        text += f"\n{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                embed.add_field(name="**점수**",
                                                                value=text,
                                                                inline=False)
                                            else:
                                                upper = ""
                                                lower = ""
                                                for i in range(4):
                                                    lv = etc.level(
                                                        q.readXpById(player[i]['id']))
                                                    if i == 0:
                                                        upper += f"{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                    else:
                                                        upper += f"\n{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                for i in range(4, len(player)):
                                                    lv = etc.level(
                                                        q.readXpById(player[i]['id']))
                                                    if i == 4:
                                                        lower += f"{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                    else:
                                                        lower += f"\n{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                embed.add_field(name="**점수**",
                                                            value=upper,
                                                            inline=False)
                                                embed.add_field(name="",
                                                            value=lower,
                                                            inline=False)

                                            embed.set_footer(
                                                text=f"{name} | CHAIN: {chain}")
                                            await ctx.send(embed=embed)
                                            chain += 1
                                            break

                                        else:
                                            player[i]['life'] -= 1
                                            player[i]['score'] -= 30
                                            await ctx.send(
                                                f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** 없는 단어입니다...'
                                            )
                                            shCheck = True
                                            break

                                    elif check[0] != start:
                                        player[i]['life'] -= 1
                                        player[i]['score'] -= 30
                                        await ctx.send(
                                            f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** **`{start}`**(으)로 시작하는 단어를 입력해 주세요...'
                                        )

                                    elif check in history:
                                        player[i]['life'] -= 1
                                        player[i]['score'] -= 50
                                        await ctx.send(
                                            f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -50점** 이미 사용한 단어 입니다...'
                                        )
                                        shCheck = True
                                        break

                                    elif len(check) < 2:
                                        player[i]['life'] -= 1
                                        player[i]['score'] -= 30
                                        await ctx.send(
                                            f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** 적어도 2글자 이상 되어야 합니다...'
                                        )
                                        shCheck = True
                                        break

                            if player[i]['life'] == 0:
                                await ctx.send(
                                    f'`(⩌ʌ ⩌;)` **{q.readTagById(uid)}** 님이 탈락하였습니다...'
                                )
                                end = True
                                break

                            elif player[i]['life'] > 0 and shCheck:
                                shCheck = False
                                player = shufflePlayer(player, i)
                                bonus = wd.random_korean()
                                print(player)
                                break


                        if end:
                            record = datetime.datetime.now().timestamp(
                            ) - start_time
                            recordt = int(record * 100)
                            embed = discord.Embed(
                                title='결과',
                                description=
                                f'CHAIN: {chain-1}\nTIME: {recordt//6000}분 {(recordt%6000)//100:02d}초 {recordt%100:02d}',
                                color=0xBCE29E)
                            
                            for i in range(len(player)):
                                if player[i]['life'] == 3:
                                    player[i]['score'] = int(player[i]['score']*2.4)

                            player.sort(key=lambda x: -x['score'])

                            for i in range(len(player)):
                                if player[i]['score'] < 0:
                                    player[i]['score'] = 0
                                xp_gain = int(
                                    (player[i]['score'] * 1.8 + chain * 5) *
                                    (1 - 0.15 * i))
                                money_gain = int(
                                    (player[i]['score'] * 1.2 + chain * 3) *
                                    (1 - 0.15 * i))
                                q.xpAddById(player[i]['id'], xp_gain)
                                q.moneyAddById(player[i]['id'], money_gain)
                                lv = etc.level(q.readXpById(player[i]['id']))
                                if i == 0:
                                    l.wcUpdateIndi(player[i]['id'],
                                                    player[i]['score'], chain - 1,
                                                    True)
                                else:
                                    l.wcUpdateIndi(player[i]['id'],
                                                    player[i]['score'], chain - 1,
                                                    False)
                                embed.add_field(
                                    name=
                                    f"{scoreFont(i+1, 1, 0)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}",
                                    value=
                                    f"{scoreFont(player[i]['score'], 4, player[i]['color'])} | {lifeUI(player[i]['life'],3)} | +{xp_gain}XP, +${money_gain}",
                                    inline=False)

                                #도전과제
                                if l.wcReadById(player[i]['id'],
                                                'win') >= 84 and q.readStorageById(
                                                    player[i]['id'], 81) == 0:
                                    q.storageModifyById(player[i]['id'], 81, 1)

                                if chain > 421 and q.readStorageById(
                                        player[i]['id'], 83) == 0:
                                    q.storageModifyById(player[i]['id'], 83, 1)

                                if l.wcReadById(
                                        player[i]['id'],
                                        'regist') >= 1446 and q.readStorageById(
                                            player[i]['id'], 84) == 0:
                                    q.storageModifyById(player[i]['id'], 84, 1)

                                if player[i]['score'] >= 1000 and q.readStorageById(
                                            player[i]['id'], 150) == 0:
                                    q.storageModifyById(player[i]['id'], 150, 1)

                                if i == 0:
                                    if player[i]['life'] == 3 and q.readStorageById(player[i]['id'], 148) == 0:
                                        q.storageModifyById(player[i]['id'], 148, 1)

                                    elif player[i]['life'] == 0 and q.readStorageById(player[i]['id'], 149) == 0:
                                        q.storageModifyById(player[i]['id'], 149, 1)


                            embed.set_footer(text='Discord Bot by Dizzt')
                            await ctx.send("## 게임 끝", embed=embed)
                            break

                elif gamestart and option == '쿵쿵따':
                    random.shuffle(player)
                    chain = 1
                    history = []
                    sample = sampleText()
                    start = random.choice(sample)
                    start_alter = ""
                    length = random.randint(2,5)
                    end = False
                    shCheck = False

                    color_arr = [1, 2, 3, 4, 5 ,6, 7, 8]
                    random.shuffle(color_arr)

                    for i in range(len(player)):
                        player[i]['color'] = color_arr[i]

                    embed = discord.Embed(title='참가자 목록',
                                            description=f'인원: {len(player)}/6',
                                            color=0xBCE29E)
                    for i in range(len(player)):
                        lv = etc.level(q.readXpById(player[i]['id']))
                        embed.add_field(
                            name=
                            f"{player_badge[player[i]['color']]}{etc.lvicon(lv)}{q.readTagById(player[i]['id'])} (Lv. {lv})",
                            value=
                            f"**{scoreFont(player[i]['score'], 4, player[i]['color'])}점** | {lifeUI(player[i]['life'],3)}",
                            inline=False)
                    embed.set_footer(text='Discord Bot by Dizzt')
                    await ctx.send(
                        f"**잠시후 게임이 시작됩니다!**\n종목: 끝말잇기 (**`{sample}`** 중 한 글자가 랜덤으로 배치 됩니다.)"
                    )
                    await asyncio.sleep(5)
                    start_time = datetime.datetime.now().timestamp()

                    while True:
                        for i in range(len(player)):

                            uid = player[i]['id']
                            ulv = etc.level(q.readXpById(uid))

                            if player[i]['life'] > 0 and isOneKill(start):
                                s = int(player[i]['score'] * 0.42)
                                player[i]['score'] -= s
                                player[i]['life'] -= 1
                                start = random.choice(sample)
                                await ctx.send(
                                    f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -{s}점** 한방 단어 공격을 받았습니다...'
                                )
                                player = shufflePlayer(player, i)
                                length = random.randint(2,5)
                                break

                            while player[i]['life'] > 0:

                                if replace_sound_char(start) is not None:
                                    start_alter = replace_sound_char(start)
                                    await ctx.send(
                                        f":chains:{scoreFont(chain, 3, 0)} | {player_badge[player[i]['color']]}{etc.lvicon(ulv)}{q.readTagById(uid)} | {scoreFont(player[i]['score'], 4, player[i]['color'])} | {lifeUI(player[i]['life'],3)} <@{uid}>\n## {start}({start_alter})\n`제한 길이` {length}글자\n(으)로 시작하는 단어를 입력하세요! ('q' 입력시 포기)"
                                    )
                                else:
                                    start_alter = ""
                                    await ctx.send(
                                        f":chains:{scoreFont(chain, 3, 0)} | {player_badge[player[i]['color']]}{etc.lvicon(ulv)}{q.readTagById(uid)} | {scoreFont(player[i]['score'], 4, player[i]['color'])} | {lifeUI(player[i]['life'],3)} <@{uid}>\n## {start}\n`제한 길이` {length}글자\n(으)로 시작하는 단어를 입력하세요! ('q' 입력시 포기)"
                                    )

                                if player[i]['id'] == 691455977270149171:
                                    if sample == '●▅▇█▇▆▅▄▇':
                                        alterlist = "가나다라마사바아자파카타파하"
                                        pick = random.randint(1, len(alterlist))
                                        start = alterlist[pick]

                                    dice = 0
                                    if detectZwong(i, player):
                                        dice = random.randint(1, 2)
                                    else:
                                        dice = random.randint(1, 7)
                                    start_alter = replace_sound_char(start)
                                    if dice == 1:
                                        result = searchKiller(start, length)
                                        if start_alter is not None:
                                            result += searchKiller(start_alter, length)
                                        if len(result) != 0:
                                            input_word = await ctx.send(random.choice(result))
                                        else:
                                            result = await asyncio.to_thread(
                                                ks.startWord,
                                                start,
                                                history,
                                                fixed_length=length,
                                            )
                                            if result is not None:
                                                input_word = await ctx.send(result[0])
                                            else:
                                                input_word = await ctx.send("q")
                                    else:
                                        result = await asyncio.to_thread(
                                            ks.startWord,
                                            start,
                                            history,
                                            fixed_length=length,
                                        )
                                        if result is not None:
                                            input_word = await ctx.send(result[0])
                                        else:
                                            input_word = await ctx.send("q")

                                else:

                                    def check(m):
                                        return m.author.id == uid and m.channel == ctx.channel

                                    input_word = await self.client.wait_for(
                                        "message", check=check)

                                check = input_word.content

                                if check == 'q':
                                    s = int(player[i]['score'] * 0.33)
                                    player[i]['score'] -= s
                                    player[i]['life'] -= 1
                                    start = random.choice(sample)
                                    await ctx.send(
                                        f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -{s}점** 방어에 실패하였습니다...'
                                    )
                                    shCheck = True
                                    break

                                else:
                                    if (
                                        check[0] == start
                                        or check[0] == start_alter
                                        or sample == '●▅▇█▇▆▅▄▇'
                                    ) and len(check) == length and check not in history:

                                        result = wd.readInGame(check)

                                        if result is None:
                                            result = await asyncio.to_thread(
                                                ks.searchWord, check
                                            )

                                        if result is not None:
                                            start = check[-1]
                                            history.append(check)
                                            player[i]['score'] += scoreBoost(
                                                count_break_korean(check),
                                                player[i]['life'])
                                            wd.newWordById(player[i]['id'],
                                                            str(result[0]),
                                                            str(result[1]),
                                                            str(result[2]))
                                            index = wd.findID(check)
                                            name = q.readTagById(player[i]['id'])
                                            embed = discord.Embed(
                                                title=f'{result[0]} `id: {index}`',
                                                description=
                                                f'[{result[1]}] {result[2]}',
                                                color=color[player[i]['color']])
                                            
                                            if len(player) < 5:
                                                text = ""
                                                for i in range(len(player)):
                                                    lv = etc.level(
                                                        q.readXpById(player[i]['id']))
                                                    if i == 0:
                                                        text += f"{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                    else:
                                                        text += f"\n{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                embed.add_field(name="**점수**",
                                                                value=text,
                                                                inline=False)
                                            else:
                                                upper = ""
                                                lower = ""
                                                for i in range(4):
                                                    lv = etc.level(
                                                        q.readXpById(player[i]['id']))
                                                    if i == 0:
                                                        upper += f"{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                    else:
                                                        upper += f"\n{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                for i in range(4, len(player)):
                                                    lv = etc.level(
                                                        q.readXpById(player[i]['id']))
                                                    if i == 4:
                                                        lower += f"{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                    else:
                                                        lower += f"\n{player_badge[player[i]['color']]} {scoreFont(player[i]['score'], 4, player[i]['color'])} {lifeUI(player[i]['life'],3)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}"
                                                embed.add_field(name="**점수**",
                                                            value=upper,
                                                            inline=False)
                                                embed.add_field(name="",
                                                            value=lower,
                                                            inline=False)

                                            embed.set_footer(
                                                text=f"{name} | CHAIN: {chain}")
                                            await ctx.send(embed=embed)
                                            chain += 1
                                            break

                                        else:
                                            player[i]['life'] -= 1
                                            player[i]['score'] -= 30
                                            await ctx.send(
                                                f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** 없는 단어입니다...'
                                            )
                                            shCheck = True
                                            break

                                    elif check[0] != start:
                                        player[i]['life'] -= 1
                                        player[i]['score'] -= 30
                                        await ctx.send(
                                            f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** **`{start}`**(으)로 시작하는 단어를 입력해 주세요...'
                                        )

                                    elif check in history:
                                        player[i]['life'] -= 1
                                        player[i]['score'] -= 50
                                        await ctx.send(
                                            f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -50점** 이미 사용한 단어 입니다...'
                                        )
                                        shCheck = True
                                        break

                                    elif len(check) != length:
                                        player[i]['life'] -= 1
                                        player[i]['score'] -= 30
                                        await ctx.send(
                                            f'`(⩌ʌ ⩌;)` <@{uid}> **-1 목숨 | -30점** {length}글자 단어만 가능합니다...'
                                        )
                                        shCheck = True
                                        break

                            if player[i]['life'] == 0:
                                await ctx.send(
                                    f'`(⩌ʌ ⩌;)` **{q.readTagById(uid)}** 님이 탈락하였습니다...'
                                )
                                end = True
                                break

                            elif player[i]['life'] > 0 and shCheck:
                                shCheck = False
                                length = random.randint(2,5)
                                player = shufflePlayer(player, i)
                                print(player)
                                break


                        if end:
                            record = datetime.datetime.now().timestamp(
                            ) - start_time
                            recordt = int(record * 100)
                            embed = discord.Embed(
                                title='결과',
                                description=
                                f'CHAIN: {chain-1}\nTIME: {recordt//6000}분 {(recordt%6000)//100:02d}초 {recordt%100:02d}',
                                color=0xBCE29E)
                            
                            for i in range(len(player)):
                                if player[i]['life'] == 3:
                                    player[i]['score'] = int(player[i]['score']*2.4)

                            player.sort(key=lambda x: -x['score'])

                            for i in range(len(player)):
                                if player[i]['score'] < 0:
                                    player[i]['score'] = 0
                                xp_gain = int(
                                    (player[i]['score'] * 1.8 + chain * 5) *
                                    (1 - 0.15 * i))
                                money_gain = int(
                                    (player[i]['score'] * 1.2 + chain * 3) *
                                    (1 - 0.15 * i))
                                q.xpAddById(player[i]['id'], xp_gain)
                                q.moneyAddById(player[i]['id'], money_gain)
                                lv = etc.level(q.readXpById(player[i]['id']))
                                if i == 0:
                                    l.wcUpdateIndi(player[i]['id'],
                                                    player[i]['score'], chain - 1,
                                                    True)
                                else:
                                    l.wcUpdateIndi(player[i]['id'],
                                                    player[i]['score'], chain - 1,
                                                    False)
                                embed.add_field(
                                    name=
                                    f"{scoreFont(i+1, 1, 0)} {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}",
                                    value=
                                    f"{scoreFont(player[i]['score'], 4, player[i]['color'])} | {lifeUI(player[i]['life'],3)} | +{xp_gain}XP, +${money_gain}",
                                    inline=False)

                                #도전과제
                                if l.wcReadById(player[i]['id'],
                                                'win') >= 84 and q.readStorageById(
                                                    player[i]['id'], 81) == 0:
                                    q.storageModifyById(player[i]['id'], 81, 1)

                                if chain > 421 and q.readStorageById(
                                        player[i]['id'], 83) == 0:
                                    q.storageModifyById(player[i]['id'], 83, 1)

                                if l.wcReadById(
                                        player[i]['id'],
                                        'regist') >= 1446 and q.readStorageById(
                                            player[i]['id'], 84) == 0:
                                    q.storageModifyById(player[i]['id'], 84, 1)

                                if player[i]['score'] >= 1000 and q.readStorageById(
                                            player[i]['id'], 150) == 0:
                                    q.storageModifyById(player[i]['id'], 150, 1)

                                if i == 0:
                                    if player[i]['life'] == 3 and q.readStorageById(player[i]['id'], 148) == 0:
                                        q.storageModifyById(player[i]['id'], 148, 1)
                                        
                                    elif player[i]['life'] == 0 and q.readStorageById(player[i]['id'], 149) == 0:
                                        q.storageModifyById(player[i]['id'], 149, 1)

                            embed.set_footer(text='Discord Bot by Dizzt')
                            await ctx.send("## 게임 끝", embed=embed)
                            break
        
        if team == '팀전':
            if option in ['일반', '쿵쿵따']:
                gamestart = False

                not_include = []
                player = []

                owner, _ = await _prepare_wordchain_participant(ctx)
                player.append({"id": owner.id, "score": 0, "life": 3, "color": 0, "team": None})
                not_include.append(owner.id)

                while True:
                    embed = discord.Embed(title='참가자 목록',
                                            description=f'인원: {len(player)}/8',
                                            color=0xBCE29E)
                    for i in range(len(player)):
                        lv = etc.level(q.readXpById(player[i]['id']))
                        sts = l.wcReadById(player[i]['id'], 'stats')
                        print(sts)
                        embed.add_field(
                            name=
                            f"`{str(player[i]['team']) + '팀' if player[i]['team'] else '자동 배정'}` {etc.lvicon(lv)}{q.readTagById(player[i]['id'])}",
                            value=f"`전적` {sts[2]}전 {sts[3]}승 | {sts[0]}점 | {sts[1]}체인",
                            inline=False)
                    embed.set_footer(text='Discord Bot by Dizzt')
                    await ctx.reply(
                        f"## 끝말잇기 ({option}) 팀전 - 인원 모집\n"
                        "* 총 2명, 4명, 6명, 8명으로 시작할 수 있습니다.\n"
                        "* `@사용자 1팀` 또는 `@사용자 2팀`으로 팀을 지정해 초대할 수 있습니다.\n"
                        "* 팀을 생략한 참가자는 시작할 때 빈자리에 자동 배정됩니다.\n"
                        "* 참가자는 `1팀` 또는 `2팀`만 입력해 자신의 팀을 직접 선택할 수 있습니다.\n"
                        "* 초대가 완료되면 `시작`, 취소하려면 `취소`를 입력해 주세요.",
                        embed=embed)

                    def check(m):
                        return (
                            m.channel == ctx.channel
                            and (m.author == ctx.author or m.author.id in not_include)
                        )

                    input_word = await self.client.wait_for("message", check=check)
                    check = input_word.content
                    actor_id = input_word.author.id

                    if check == '시작':
                        if actor_id != ctx.author.id:
                            await ctx.send("`(⩌ʌ ⩌;)` 게임은 방장만 시작할 수 있습니다.")
                            continue
                        if len(player) in (2, 4, 6, 8):
                            try:
                                _balance_wordchain_teams(player)
                            except ValueError:
                                await ctx.send(
                                    '`(⩌ʌ ⩌;)` 한 팀에 지정된 인원이 너무 많습니다. 팀 선택을 다시 확인해 주세요.')
                            else:
                                gamestart = True
                                await ctx.send(
                                    ":green_circle: 두 팀이 같은 인원으로 구성되었습니다. 게임을 시작합니다!"
                                )
                                break
                        else:
                            await ctx.send(
                                '`(⩌ʌ ⩌;)` 팀전은 총 2명, 4명, 6명, 8명일 때만 시작할 수 있습니다.')

                    elif check == '취소':
                        if actor_id != ctx.author.id:
                            await ctx.send("`(⩌ʌ ⩌;)` 게임은 방장만 취소할 수 있습니다.")
                            continue
                        await ctx.send(":x: 게임 생성이 취소되었습니다.")
                        break

                    else:
                        try:
                            reference, selected_team = _parse_team_selection(check)
                            if reference is None:
                                existing = next(
                                    item for item in player if item["id"] == actor_id
                                )
                                existing["team"] = selected_team
                                await ctx.send(
                                    f":green_circle: `{q.readTagById(actor_id)}` 님이 {selected_team}팀을 선택했습니다."
                                )
                                continue

                            if actor_id != ctx.author.id:
                                await ctx.send(
                                    "`(⩌ʌ ⩌;)` 다른 참가자의 초대와 팀 변경은 방장만 할 수 있습니다."
                                )
                                continue

                            invited, _ = await _prepare_wordchain_participant(ctx, reference)
                            id = invited.id
                            if id in not_include:
                                existing = next(item for item in player if item["id"] == id)
                                if selected_team is not None:
                                    existing["team"] = selected_team
                                    await ctx.send(
                                        f":green_circle: `{q.readTagById(id)}` 님을 {selected_team}팀으로 변경했습니다."
                                    )
                                else:
                                    await ctx.send(
                                    '`(⩌ʌ ⩌;)` 이미 참가한 사람입니다...')
                            elif len(player) >= 8:
                                await ctx.send(
                                    '`(⩌ʌ ⩌;)` 인원이 너무 많습니다... 최대 8명 까지 참가가 가능합니다!')
                            else:
                                name = q.readTagById(id)
                                player.append({
                                    "id": id,
                                    "score": 0,
                                    "life": 3,
                                    "color": 0,
                                    "team": selected_team,
                                })
                                not_include.append(id)
                                team_label = f" ({selected_team}팀)" if selected_team else " (자동 배정)"
                                await ctx.send(
                                    f":green_circle: `{name}`가 참가자 목록에 추가되었습니다!{team_label}")
                        except (UserResolutionError, ValueError):
                            await ctx.send(
                                '`(⩌ʌ ⩌;)` 유효하지 않은 참가자 입니다... 다시 시도해 보세요...')

                if gamestart:
                    await _run_team_wordchain(self.client, ctx, player, option)

    # Wordchain Test [ID: 48]
    @commands.hybrid_command(name='테스트용',
                             description="테스트")
    async def wordchain_test(self, ctx):
        wd.readPOSCount()
        await ctx.send('테스트')


async def setup(client):
    await client.add_cog(TestCommands(client))

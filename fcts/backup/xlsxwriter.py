# 0. 모듈 불러오기
from openpyxl import load_workbook, Workbook
import random as r
import functions.sqlcontrol as q

# 1. 기본 설정

# 1.1. 기본 머릿글 설정
c_id = 1
c_discrim = 2
c_nick = 3
c_xp = 4
c_money = 5
c_skin = 6
c_storage = 7

opt_arr = dict(id=1, discrim=2, nick=3, xp=4, money=5, skin=6, storage=7)

# 1.2. 작업 시트 불러오기
wb = load_workbook("userdb.xlsx")
ws = wb['main']
wst = wb['storage']


def loadFile():
    wb = load_workbook("userdb.xlsx")
    ws = wb['main']
    wst = wb['storage']


def saveFile():
    wb.save("userdb.xlsx")
    wb.close()


# 1.3. 레벨 경험치 데이터

# 1.3.1 최고 레벨
final_lv = 240

# 1.3.2 앰블럼 이름 데이터
arr_emblem = [[
    'Initial', 'Scarlet', 'Amber', 'Topaz', 'Chlorophyll', 'Aquatic',
    'Sapphire', 'Violette', 'Cosmic', 'Ultimate'
],
              [
                  'Square', 'Ring', 'Trigon', 'Stelle', 'Hexagon', 'Diamond',
                  'Hexagram', 'Insignia', 'Plasma', 'Collar', 'Cardioid',
                  'Fluid', 'Flame', 'Bolt', 'Solarus', 'Polyphemus', 'Spiral',
                  'Physis', 'Elementum', 'Nebula', 'Aeternum', 'Libertas',
                  'Iustitia', 'Transcendence'
              ]]


# 1.3.3 경험치 데이터
def xpList():
    file = open("./config/xp.txt", "r")
    arr = []
    line = file.readline()
    arr.append(line.rstrip('\n'))
    while line:
        try:
            line = file.readline().rstrip('\n')
            arr.append(int(line))
        except:
            break
    file.close()
    return arr


xp_arr = xpList()
max_xp = xp_arr[final_lv - 1]

# 2. 기능 보조 함수


# 2.1. 빈 줄 찾기
def checkRow():
    loadFile()
    for row in range(2, ws.max_row + 2):
        if ws.cell(row, 1).value is None:
            return row
            break
    saveFile()
    #return ws.max_row 불안정하지만 훨씬 간단함


# 2.2. 유저 데이터 기록 여부 확인
def checkId(uid):
    sid = str(uid)
    loadFile()

    result = False

    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=c_id).value == sid:
            result = True
            break

    saveFile()

    return result


# 2.3. 경험치 관련


# 2.3.1 최고 레벨 출력
def maxLevel():
    return final_lv


# 2.3.2 정수 레벨 출력
def level(exp):
    if exp > max_xp:
        return final_lv
    else:
        i = 1
        while (1):
            if exp < xp_arr[i]:
                break
            else:
                i += 1
        return i


# 2.3.3. Required XP
def need_exp(i):
    return int(xp_arr[i])


# 2.3.4. Level Up (Boolean)
def level_up(temp, pres):
    if temp > final_lv:
        return False
    else:
        if need_exp(temp) < pres:
            return True
        else:
            return False


# 2.3.5. emblem name
def emblemName(lv):
    return "{} {} Emblem".format(arr_emblem[0][int((lv - 1) % 10)],
                                 arr_emblem[1][int((lv - 1) / 10)])


# 3. 주 기능 함수


# 3.1 초기 설정 함수
def signUp(uid, name, xp_value, money_value):

    loadFile()

    #빈 줄 찾기
    empty_row = checkRow()

    discrim = r.randint(1, 10000)
    init_skin = 1
    init_storage = int(empty_row)
    sid = str(uid)

    ws.cell(row=empty_row, column=c_id, value=sid)
    ws.cell(row=empty_row, column=c_discrim, value=discrim)
    ws.cell(row=empty_row, column=c_nick, value=name)
    ws.cell(row=empty_row, column=c_xp, value=xp_value)
    ws.cell(row=empty_row, column=c_money, value=money_value)
    ws.cell(row=empty_row, column=c_skin, value=init_skin)
    ws.cell(row=empty_row, column=c_storage, value=init_storage)

    wst.cell(row=empty_row, column=1, value=sid)
    wst.cell(row=empty_row, column=2, value=1)
    for i in range(0, 255):
        wst.cell(row=empty_row, column=i + 3, value=0)

    saveFile()


# 3.2 데이터 수정


# 3.2.1. 통으로 수정하기
def dataWrite(uid, option, value):
    sid = str(uid)
    loadFile()

    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=c_id).value == sid:
            ws.cell(row=row, column=opt_arr[option], value=value)
            break

    saveFile()


# 3.2.2. 증감 연산 수행
def dataAddup(uid, option, amount):
    sid = str(uid)
    loadFile()
    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=c_id).value == sid:
            if option == 'xp':
                ws.cell(row, c_xp).value += amount
            elif option == 'money':
                ws.cell(row, c_money).value += amount
            break
    saveFile()


# 3.2.3. 총원 증감 연산 수행
def dataAddupAll(option, amount):
    loadFile()
    for row in range(2, ws.max_row + 1):
        if option == 'xp':
            ws.cell(row, c_xp).value += amount
        elif option == 'money':
            ws.cell(row, c_money).value += amount
    saveFile()


# 3.2.4. 스킨 데이터 수정
def storageWrite(uid, row, skin_id, mark):
    loadFile()
    if str(uid) == wst.cell(row, 1).value:
        if mark == 1 and wst.cell(row, skin_id + 1).value == 0:
            wst.cell(row, skin_id + 1).value = 1
        elif mark == 0 and wst.cell(row, skin_id + 1).value == 1:
            wst.cell(row, skin_id + 1).value = 0
    saveFile()


# 3.3. 데이터 불러오기


# 3.3.1. 데이터 낱개로 불러오기
def dataRead(uid, option):

    check_row = 2
    sid = str(uid)

    loadFile()

    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=c_id).value == sid:
            check_row = row
            break

    result = ws.cell(check_row, opt_arr[option]).value

    saveFile()

    if option == 'id':
        return int(result)
    elif option == 'discrim':
        return str(result).zfill(4)
    else:
        return result


# 3.3.2. 데이터 묶음으로 불러오기
def dataReadAll(uid):
    check_row = 2
    sid = str(uid)

    loadFile()
    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=c_id).value == sid:
            check_row = row
            break

    result = {
        'id': ws.cell(check_row, c_id).value,
        'discrim': str(ws.cell(check_row, c_discrim).value).zfill(4),
        'nick': ws.cell(check_row, c_nick).value,
        'xp': ws.cell(check_row, c_xp).value,
        'money': ws.cell(check_row, c_money).value,
        'skin': ws.cell(check_row, c_skin).value,
        'storage': ws.cell(check_row, c_storage).value
    }

    saveFile()

    return result


# 3.3.3. 스킨 데이터 낱개로 불러오기
def storageRead(uid, row, skin_id):
    result = 0
    loadFile()
    if str(uid) == wst.cell(row, 1).value:
        result = wst.cell(row, skin_id + 1).value
    saveFile()
    return result


def storageRead2():
    result = 0
    users = q.idList()
    print(users)

    loadFile()
    for u in users:
        try:
            q.newStorageById(int(u[0]))
            print(u[0])
        except:
            print(str(u[0]) + ":pass")
        for row in range(2, ws.max_row + 2):
            if wst.cell(row=row, column=1).value == str(u[0]):
                for i in range(1, 30):
                    value = wst.cell(row, i + 1).value
                    if value == 1:
                        q.storageModifyById(int(u[0]), str(i), 1)
    saveFile()
    return result
